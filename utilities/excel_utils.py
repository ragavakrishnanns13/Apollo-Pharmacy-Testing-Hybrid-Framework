from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class ExcelReader:
    """
    Class Name: ExcelReader
    Author: Srishti Kundu
    Description: Reads data from the Excel file, when we pass row number and column name
    Return Type: String
    Parameters: Row Number, Column Name
    """
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path

    def get_data(self, sheet_name=None, sheet_number=None):
        wb = load_workbook(self.excel_file_path, data_only=True)
        if sheet_name is not None:
            ws = wb[sheet_name]
        elif sheet_number is not None:
            ws = wb.worksheets[sheet_number]
        else:
            raise ValueError("Either sheet_name or sheet_number must be provided")
        data = []
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        return data

    # def read_data(self, sheet_name, row_number, column_name):
    #     wb = load_workbook(self.excel_file_path, data_only=True)
    #     ws = wb[sheet_name]
    #     headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    #     col_idx = headers.index(column_name) + 1
    #     cell_value = ws.cell(row=row_number + 1, column=col_idx).value
    #     return cell_value

    def read_data(self, sheet_name, row_number, column_name):
        wb = load_workbook(self.excel_file_path, data_only=True)
        ws = wb[sheet_name]

        try:
            # Try treating column_name as a header
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            col_idx = headers.index(column_name) + 1
        except ValueError:
            # If not found, treat it as a column letter
            col_idx = column_index_from_string(column_name)

        cell_value = ws.cell(row=row_number, column=col_idx).value
        return cell_value

    def write_data(self, sheet_name, row_number, column_name, value):
        wb = load_workbook(self.excel_file_path)
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = headers.index(column_name) + 1
        ws.cell(row=row_number + 1, column=col_idx, value=value)
        wb.save(self.excel_file_path)

# USAGE
# self.reader = ExcelReader("Testdata/excel.xlsx") #put the file relevant filepath in this and put it in teh init constructor of the class
# valuepos = self.reader.read_data(sheet_name="Sheet1", row_number=1, column_name="JOB TITLE")
# self.helper.sendKeys(CareerLocator.JOB_INPUT, valuepos)